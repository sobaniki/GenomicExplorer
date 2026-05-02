#!/usr/bin/env python3

"""Preprocess: Split (group-wise split)

This plugin splits an *already matched* phenotype table (typically phenotype.tsv
created by prep_split_merge_extract) into group-wise files using a specified
column (e.g., Panel).

This plugin splits an *already matched* phenotype table into group-wise files
using a specified column (e.g., Panel). Optionally, it can also split an
*already matched* genotype file (VCF or PLINK bed/bim/fam) using the generated
per-group sample lists.

IMPORTANT
---------
This plugin assumes genotype sample IDs are already aligned to phenotype sample
IDs (e.g., after running Split/Merge/Extract + Extract). It will NOT attempt to
rename or match IDs.

Outputs:
- groups/<group_col>/<group>/phenotype.tsv
- groups/<group_col>/<group>/selected_samples.txt (optional)
- groups/<group_col>/<group>/keep.fid_iid.tsv (optional; FID=0)
- groups/<group_col>/<group>/genotype/*  (optional)
- split_summary.tsv

Inputs (params.json)
--------------------
out_dir_override: str (optional)
phenotype_path: str (required; tsv/csv/xlsx)
phenotype_sheet: str (optional; xlsx only)
sample_id_col: str (optional; default 'id' or first column)
group_col: str (required)
group_values: str (optional; comma-separated values; if empty, all unique)
min_samples: int (optional; default 1)
write_sample_list: bool (default True)
write_keep_fid_iid: bool (default True)
genotype_mode: str (optional; one of none|vcf|plink)
genotype_vcf_path: str (required if genotype_mode=vcf)
genotype_plink_prefix: str (required if genotype_mode=plink)
bcftools_bin: str (optional; default 'bcftools')
plink2_bin: str (optional; default 'plink2')

Safety guard:
If group_col appears to be nearly unique per row (e.g., accidentally set to the
sample ID column), the plugin aborts with a descriptive error.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import os
import subprocess
from pathlib import Path
from typing import Dict, List, Tuple, Optional


def _read_text(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="ignore")


def _write_text(p: Path, s: str):
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(s, encoding="utf-8")


def _run_cmd(cmd: List[str], *, cwd: Optional[Path] = None, stdout_path: Optional[Path] = None, stderr_path: Optional[Path] = None) -> int:
    """Run external command, optionally logging stdout/stderr to files."""
    if stdout_path is not None:
        stdout_path.parent.mkdir(parents=True, exist_ok=True)
    if stderr_path is not None:
        stderr_path.parent.mkdir(parents=True, exist_ok=True)
    out_f = stdout_path.open("w", encoding="utf-8") if stdout_path else open(os.devnull, "w", encoding="utf-8")
    err_f = stderr_path.open("w", encoding="utf-8") if stderr_path else open(os.devnull, "w", encoding="utf-8")
    try:
        r = subprocess.run(cmd, cwd=str(cwd) if cwd else None, stdout=out_f, stderr=err_f)
        return int(r.returncode)
    finally:
        try:
            out_f.close()
        except Exception:
            pass
        try:
            err_f.close()
        except Exception:
            pass


def _run_bcftools_view(
    bcftools_bin: str,
    in_vcf: Path,
    keep_samples: Path,
    out_vcf: Path,
    *,
    log_dir: Path,
) -> int:
    """Try bcftools view -S on a file. Returns exit code."""
    cmd = [bcftools_bin, "view", "--force-samples", "-S", str(keep_samples), "-Oz", "-o", str(out_vcf), str(in_vcf)]
    rc = _run_cmd(cmd, stdout_path=log_dir / "bcftools_stdout.txt", stderr_path=log_dir / "bcftools_stderr.txt")
    if rc == 0:
        # index (optional)
        _run_cmd([bcftools_bin, "index", "-f", "-t", str(out_vcf)], stdout_path=log_dir / "bcftools_index_stdout.txt", stderr_path=log_dir / "bcftools_index_stderr.txt")
    return rc


def _run_bcftools_view_sanitized(
    bcftools_bin: str,
    in_vcf: Path,
    keep_samples: Path,
    out_vcf: Path,
    *,
    log_dir: Path,
) -> int:
    """Fallback: stream VCF through grep -v '^##SAMPLE=<' then bcftools view."""
    log_dir.mkdir(parents=True, exist_ok=True)
    err_path = log_dir / "bcftools_sanitize_stderr.txt"
    out_log = log_dir / "bcftools_sanitize_stdout.txt"

    cmd_in = ["zcat", str(in_vcf)] if in_vcf.suffix.lower().endswith("gz") else ["cat", str(in_vcf)]
    cmd_grep = ["grep", "-v", r"^##SAMPLE=<"]
    cmd_bcft = [bcftools_bin, "view", "--force-samples", "-S", str(keep_samples), "-Oz", "-o", str(out_vcf), "-"]

    with out_log.open("w", encoding="utf-8") as out_f, err_path.open("w", encoding="utf-8") as err_f:
        p1 = subprocess.Popen(cmd_in, stdout=subprocess.PIPE, stderr=err_f)
        p2 = subprocess.Popen(cmd_grep, stdin=p1.stdout, stdout=subprocess.PIPE, stderr=err_f)
        assert p1.stdout is not None
        p1.stdout.close()
        p3 = subprocess.Popen(cmd_bcft, stdin=p2.stdout, stdout=subprocess.DEVNULL, stderr=err_f)
        assert p2.stdout is not None
        p2.stdout.close()
        rc3 = int(p3.wait())
        # also wait upstream
        _ = p2.wait()
        _ = p1.wait()
        out_f.write(" ".join(cmd_in) + "\n")
        out_f.write(" ".join(cmd_grep) + "\n")
        out_f.write(" ".join(cmd_bcft) + "\n")

    if rc3 == 0:
        _run_cmd([bcftools_bin, "index", "-f", "-t", str(out_vcf)], stdout_path=log_dir / "bcftools_sanitize_index_stdout.txt", stderr_path=log_dir / "bcftools_sanitize_index_stderr.txt")
    return rc3


def _run_plink2_split_vcf(
    plink2_bin: str,
    in_vcf: Path,
    keep_fid_iid: Path,
    out_prefix: Path,
    *,
    log_dir: Path,
) -> int:
    cmd = [
        plink2_bin,
        "--vcf",
        str(in_vcf),
        "--keep",
        str(keep_fid_iid),
        "--export",
        "vcf",
        "bgz",
        "--out",
        str(out_prefix),
    ]
    return _run_cmd(cmd, stdout_path=log_dir / "plink2_export_vcf_stdout.txt", stderr_path=log_dir / "plink2_export_vcf_stderr.txt")


def _run_plink2_split_bed(
    plink2_bin: str,
    bfile_prefix: Path,
    keep_fid_iid: Path,
    out_prefix: Path,
    *,
    log_dir: Path,
) -> int:
    cmd = [
        plink2_bin,
        "--bfile",
        str(bfile_prefix),
        "--keep",
        str(keep_fid_iid),
        "--make-bed",
        "--out",
        str(out_prefix),
    ]
    return _run_cmd(cmd, stdout_path=log_dir / "plink2_make_bed_stdout.txt", stderr_path=log_dir / "plink2_make_bed_stderr.txt")


def _as_bool(v, default: bool = False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("1", "true", "t", "yes", "y", "on"):
        return True
    if s in ("0", "false", "f", "no", "n", "off"):
        return False
    return default


def _sanitize_group_name(name: str) -> str:
    s = (name or "").strip()
    if not s:
        return "NA"
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_.-]+", "_", s)
    s = s.strip("._-")
    return s or "NA"


def _read_table(path: Path, sheet: str = "") -> Tuple[List[str], List[Dict[str, str]]]:
    ext = path.suffix.lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise SystemExit(f"openpyxl is required to read {ext} files: {e}")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = None
        out_rows: List[Dict[str, str]] = []
        for r in rows:
            if header is None:
                header = [str(x).strip() if x is not None else "" for x in r]
                header = [h if h else f"V{idx+1}" for idx, h in enumerate(header)]
                continue
            if r is None:
                continue
            rec = {}
            for i, col in enumerate(header):
                v = r[i] if i < len(r) else None
                rec[col] = "" if v is None else str(v)
            out_rows.append(rec)
        if not header:
            raise SystemExit(f"No header row detected in spreadsheet: {path}")
        return header, out_rows

    # TSV/CSV
    # auto-detect delimiter by extension; allow .txt as TSV
    delim = "\t" if ext in (".tsv", ".txt") else ","
    text = _read_text(path)
    # allow fallback if commas exist in tsv (rare)
    if delim == "\t" and "\t" not in text and "," in text:
        delim = ","

    lines = text.splitlines()
    if not lines:
        raise SystemExit(f"Empty file: {path}")

    reader = csv.DictReader(lines, delimiter=delim)
    header = list(reader.fieldnames or [])
    if not header:
        raise SystemExit(f"Failed to read header: {path}")
    out_rows = []
    for row in reader:
        if row is None:
            continue
        out_rows.append({k: ("" if row.get(k) is None else str(row.get(k))) for k in header})
    return header, out_rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--params", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    params_path = Path(args.params)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    p = json.loads(_read_text(params_path) or "{}")

    out_override = str(p.get("out_dir_override") or "").strip()
    root = Path(out_override).expanduser().resolve() if out_override else out_dir
    root.mkdir(parents=True, exist_ok=True)

    phenotype_path = Path(str(p.get("phenotype_path") or "").strip()).expanduser()
    if not phenotype_path.exists():
        raise SystemExit(f"phenotype_path not found: {phenotype_path}")

    sheet = str(p.get("phenotype_sheet") or "").strip()
    sample_id_col = str(p.get("sample_id_col") or "").strip()
    group_col = str(p.get("group_col") or "").strip()
    if not group_col:
        raise SystemExit("group_col is required")

    group_values_raw = str(p.get("group_values") or "").strip()
    group_values = [v.strip() for v in re.split(r"[,;\s]+", group_values_raw) if v.strip()] if group_values_raw else []

    min_samples = int(p.get("min_samples") or 1)
    min_samples = max(1, min_samples)
    # Always write group-wise sample list and keep files (GUI defaults).
    write_sample_list = True
    write_keep = True

    # Optional genotype split (after Extract, genotype sample IDs must match phenotype IDs)
    split_genotype = _as_bool(p.get('split_genotype') or p.get('do_genotype_split') or p.get('genotype_split'), False)
    genotype_path_s = str(p.get('genotype_path') or p.get('genotype_file') or '').strip()
    genotype_out_prefix = str(p.get('genotype_out_prefix') or 'genotype').strip() or 'genotype'

    bcftools_bin = str(p.get('bcftools_bin') or 'bcftools').strip() or 'bcftools'
    plink2_bin = str(p.get('plink2_bin') or 'plink2').strip() or 'plink2'

    genotype_mode = 'none'
    genotype_input = None
    if split_genotype:
        if not genotype_path_s:
            raise SystemExit('split_genotype is enabled but genotype_path is empty')
        gp = Path(genotype_path_s).expanduser()
        if gp.exists() and (gp.name.lower().endswith('.vcf') or gp.name.lower().endswith('.vcf.gz') or gp.name.lower().endswith('.bcf') or gp.name.lower().endswith('.bcf.gz')):
            genotype_mode = 'vcf'
            genotype_input = gp
        else:
            # treat as PLINK prefix (accept .bed/.bim/.fam or .pgen/.pvar/.psam)
            if gp.suffix.lower() in ('.bed', '.bim', '.fam', '.pgen', '.pvar', '.psam'):
                gp = gp.with_suffix('')
            if not ((Path(str(gp) + '.bed').exists() and Path(str(gp) + '.bim').exists() and Path(str(gp) + '.fam').exists()) or (Path(str(gp) + '.pgen').exists() and Path(str(gp) + '.pvar').exists() and Path(str(gp) + '.psam').exists())):
                raise SystemExit(f'Genotype input not found / unsupported: {genotype_path_s}')
            genotype_mode = 'plink'
            genotype_input = gp

    header, rows = _read_table(phenotype_path, sheet=sheet)
    if not rows:
        raise SystemExit("phenotype table has no rows")

    # pick id column
    if sample_id_col and sample_id_col in header:
        id_col = sample_id_col
    elif "id" in header:
        id_col = "id"
    else:
        id_col = header[0]

    if group_col not in header:
        raise SystemExit(f"group_col '{group_col}' not found in columns: {', '.join(header[:30])}{'...' if len(header)>30 else ''}")

    # Determine candidate groups
    all_groups = [str(r.get(group_col, "") or "").strip() for r in rows]
    uniq_groups = sorted({g for g in all_groups if g != ""})

    # Guard: if group column is almost unique per row, abort (prevents per-sample split)
    n_rows = len(rows)
    n_uniq = len(uniq_groups)
    if n_uniq > 200 and (n_uniq / float(max(1, n_rows))) > 0.7:
        raise SystemExit(
            f"group_col '{group_col}' appears to be nearly unique per row (unique={n_uniq} / rows={n_rows}). "
            f"This likely indicates the wrong column (e.g., using sample ID). Please choose a grouping column like Panel/Population."
        )

    allowed = set(group_values) if group_values else None

    # Collect
    by_group: Dict[str, List[Dict[str, str]]] = {}
    for r in rows:
        g = str(r.get(group_col, "") or "").strip()
        if not g:
            continue
        if allowed is not None and g not in allowed:
            continue
        by_group.setdefault(g, []).append(r)

    if not by_group:
        msg = f"No rows matched the requested groups in column '{group_col}'."
        if allowed is not None:
            msg += f" Requested groups: {', '.join(sorted(allowed))}."
        msg += f" Available groups (first 30): {', '.join(uniq_groups[:30])}"
        raise SystemExit(msg)

    # Write
    groups_dir = root / "groups" / _sanitize_group_name(group_col)
    groups_dir.mkdir(parents=True, exist_ok=True)

    summary = []
    geno_errors: List[str] = []
    for g, rs in sorted(by_group.items(), key=lambda kv: kv[0]):
        if len(rs) < min_samples:
            continue
        gname = _sanitize_group_name(g)
        out_g = groups_dir / gname
        out_g.mkdir(parents=True, exist_ok=True)

        # phenotype.tsv
        ph_path = out_g / "phenotype.tsv"
        with ph_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f, delimiter="\t")
            w.writerow(header)
            for r in rs:
                w.writerow([r.get(c, "") for c in header])

        # ids
        ids = []
        seen = set()
        for r in rs:
            sid = str(r.get(id_col, "") or "").strip()
            if not sid:
                continue
            if sid in seen:
                continue
            seen.add(sid)
            ids.append(sid)

        sample_list_path = out_g / "selected_samples.txt"
        keep_path = out_g / "keep.fid_iid.tsv"

        if write_sample_list:
            _write_text(sample_list_path, "\n".join(ids) + "\n")

        if write_keep:
            _write_text(keep_path, "\n".join([f"0\t{sid}" for sid in ids]) + "\n")
        # Optional genotype split
        geno_out = ""
        geno_engine = ""
        if genotype_mode != "none" and ids and genotype_input is not None:
            geno_dir = out_g / "genotype"
            geno_dir.mkdir(parents=True, exist_ok=True)
            log_dir = geno_dir / "logs"
            log_dir.mkdir(parents=True, exist_ok=True)

            keep_path = out_g / "keep.fid_iid.tsv"  # always exists
            out_prefix = geno_dir / genotype_out_prefix

            if genotype_mode == 'vcf':
                cmd = [
                    plink2_bin,
                    '--vcf', str(genotype_input),
                    '--keep', str(keep_path),
                    '--export', 'vcf', 'bgz',
                    '--out', str(out_prefix),
                ]
                rc = _run_cmd(cmd, cwd=None, stdout_path=log_dir / 'plink2_stdout.txt', stderr_path=log_dir / 'plink2_stderr.txt')
                geno_engine = 'plink2_export_vcf'
                out_v = Path(str(out_prefix) + '.vcf.gz')
                if rc == 0 and out_v.exists():
                    geno_out = str(out_v)
                    try:
                        _run_cmd([bcftools_bin, 'index', '-t', str(out_v)], cwd=None, stdout_path=log_dir / 'bcftools_index_stdout.txt', stderr_path=log_dir / 'bcftools_index_stderr.txt')
                    except Exception:
                        pass
                else:
                    err_msg = f"VCF split failed for group '{g}'. See {log_dir}"
                    _write_text(geno_dir / 'genotype_split_error.txt', err_msg + "\n")
                    geno_errors.append(err_msg)

            elif genotype_mode == 'plink':
                cmd = [plink2_bin]
                if Path(str(genotype_input) + '.pgen').exists():
                    cmd += ['--pfile', str(genotype_input)]
                else:
                    cmd += ['--bfile', str(genotype_input)]
                cmd += ['--keep', str(keep_path), '--make-bed', '--out', str(out_prefix)]
                rc = _run_cmd(cmd, cwd=None, stdout_path=log_dir / 'plink2_stdout.txt', stderr_path=log_dir / 'plink2_stderr.txt')
                geno_engine = 'plink2_make_bed'
                if rc == 0 and Path(str(out_prefix) + '.bed').exists():
                    geno_out = str(out_prefix)
                else:
                    err_msg = f"PLINK split failed for group '{g}'. See {log_dir}"
                    _write_text(geno_dir / 'genotype_split_error.txt', err_msg + "\n")
                    geno_errors.append(err_msg)

        summary.append({

            "group_col": group_col,
            "group": g,
            "group_dir": str(out_g),
            "n_rows": len(rs),
            "n_unique_samples": len(ids),
            "genotype_mode": genotype_mode,
            "genotype_engine": geno_engine,
            "genotype_out": geno_out,
        })

    if not summary:
        raise SystemExit(f"All groups were filtered out by min_samples={min_samples}.")

    # split_summary.tsv
    sum_path = root / "split_summary.tsv"
    with sum_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "group_col",
                "group",
                "group_dir",
                "n_rows",
                "n_unique_samples",
                "genotype_mode",
                "genotype_engine",
                "genotype_out",
            ],
            delimiter="\t",
        )
        w.writeheader()
        for r in summary:
            w.writerow(r)

    if geno_errors:
        # Fail loud so the user notices incomplete outputs
        raise SystemExit("Genotype split failed for one or more groups. First error: " + geno_errors[0])

    artifacts = {
        "split_summary_tsv": str(sum_path),
        "groups_dir": str(groups_dir),
        "phenotype_path": str(phenotype_path),
        "group_col": group_col,
        "n_groups": len(summary),
        "genotype_mode": genotype_mode,
    }
    _write_text(root / "artifacts.json", json.dumps(artifacts, indent=2))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
